from pathlib import Path
import csv
import json
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak, KeepTogether
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
POS=('FL','FR','RL','RR')

TEAM=[
 ('Saai Varshan S','Team Lead','B.E. Mechanical Engineering · 3rd Year','Overall project, system architecture, coordination, testing and final decisions.','saaivarshan69@gmail.com','+91 93422 96487','https://www.linkedin.com/in/saai-varshan-8a62b7328'),
 ('Suheerthan S','Mechanical & Vehicle Dynamics Engineer','B.E. Mechanical Engineering · 3rd Year','Vehicle model, load distribution, suspension/load concepts, mechanical design and CAD.','suheerthan2514@gmail.com','+91 90433 09288','https://www.linkedin.com/in/suheerthan-s-18aa58327/'),
 ('Santhosh R','Electronics & Sensor Integration Engineer','B.E. Mechanical Engineering · 3rd Year','Sensors, wiring, data acquisition, calibration and hardware testing.','santhoshravijv@gmail.com','+91 97105 81763','https://www.linkedin.com/in/santhosh-ravi-95097b328')]

def generate_charts(payload,out):
    out=Path(out); out.mkdir(parents=True,exist_ok=True); rs=payload['results']; x=list(range(1,len(rs)+1)); s={p:[r['raw'][p] for r in rs] for p in POS}; paths=[]
    def save(fig,name):
        p=out/name; fig.tight_layout(); fig.savefig(p,dpi=180,bbox_inches='tight'); plt.close(fig); paths.append(p)
    # 01 — exact four-wheel trend family
    fig,ax=plt.subplots(figsize=(11,4.8));
    for p in POS: ax.plot(x,s[p],label=p,linewidth=1.7)
    ax.set(title='Four-Wheel Load Trend',xlabel='Time / Row',ylabel='Load [kg]'); ax.grid(alpha=.25); ax.legend(ncol=4); save(fig,'01_four_wheel_load_trend.png')
    # 02 — force transformation
    fig,ax=plt.subplots(figsize=(11,4.8));
    for p in POS: ax.plot(x,[v*9.80665 for v in s[p]],label=p,linewidth=1.5)
    ax.set(title='Force vs Time',xlabel='Time / Row',ylabel='Force [N]'); ax.grid(alpha=.25); ax.legend(ncol=4); save(fig,'02_force_vs_time_N.png')
    totals=[r['total'] for r in rs];
    fig,ax=plt.subplots(figsize=(11,4.8)); ax.plot(x,totals,linewidth=2); ax.set(title='Total Vehicle Load',xlabel='Time / Row',ylabel='Total Load [kg]'); ax.grid(alpha=.25); save(fig,'03_total_vehicle_load.png')
    front=[r['front_axle'] for r in rs]; rear=[r['rear_axle'] for r in rs];
    fig,ax=plt.subplots(figsize=(11,4.8)); ax.plot(x,front,label='Front Axle'); ax.plot(x,rear,label='Rear Axle'); ax.set(title='Front vs Rear Axle Load',xlabel='Time / Row',ylabel='Load [kg]'); ax.grid(alpha=.25); ax.legend(); save(fig,'04_front_rear_axle.png')
    left=[r['left_side'] for r in rs]; right=[r['right_side'] for r in rs];
    fig,ax=plt.subplots(figsize=(11,4.8)); ax.plot(x,left,label='Left Side'); ax.plot(x,right,label='Right Side'); ax.set(title='Left vs Right Load Distribution',xlabel='Time / Row',ylabel='Load [kg]'); ax.grid(alpha=.25); ax.legend(); save(fig,'05_left_right_distribution.png')
    dev=[max(abs(v) for v in r['deviations'].values()) for r in rs]; th=payload['summary']['threshold_kg'];
    fig,ax=plt.subplots(figsize=(11,4.8)); ax.plot(x,dev,label='Maximum absolute deviation'); ax.axhline(th,linestyle='--',label=f'±{th:g} kg screening'); ax.set(title='Load Imbalance / Deviation',xlabel='Time / Row',ylabel='Absolute Deviation [kg]'); ax.grid(alpha=.25); ax.legend(); save(fig,'06_deviation_threshold.png')
    counts=[sum(p in r['pre_alerts'] for r in rs) for p in POS];
    fig,ax=plt.subplots(figsize=(8,4.8)); ax.bar(POS,counts); ax.set(title='Pre-Equalization Alert Frequency',xlabel='Wheel Position',ylabel='Rows Flagged'); ax.grid(axis='y',alpha=.25); save(fig,'07_event_frequency.png')
    eqdiff=[max(abs(r['equalized'][p]-r['raw'][p]) for p in POS) for r in rs];
    fig,ax=plt.subplots(figsize=(11,4.8)); ax.plot(x,eqdiff); ax.set(title='Equalization Adjustment Magnitude',xlabel='Time / Row',ylabel='Maximum Adjustment [kg]'); ax.grid(alpha=.25); save(fig,'08_equalization_adjustment.png')
    fig=plt.figure(figsize=(11,6)); ax=fig.add_subplot(111,projection='3d'); nodes={p:i for i,p in enumerate(POS)}
    for p in POS: ax.bar3d(x,[nodes[p]]*len(x),[0]*len(x),[0.55]*len(x),[0.55]*len(x),s[p],alpha=.65)
    ax.set_title('3D Time × Node × Load Distribution'); ax.set_xlabel('Time / Row'); ax.set_ylabel('Measurement Node'); ax.set_zlabel('Load [kg]'); ax.set_yticks(range(4)); ax.set_yticklabels(POS); save(fig,'09_3d_time_node_load.png')
    # 10 — equalized trend, explicitly separate from raw trend
    fig,ax=plt.subplots(figsize=(11,4.8));
    for p in POS: ax.plot(x,[r['equalized'][p] for r in rs],label=p,linewidth=1.5)
    ax.set(title='Equalized Four-Wheel Trend',xlabel='Time / Row',ylabel='Equalized Load [kg]'); ax.grid(alpha=.25); ax.legend(ncol=4); save(fig,'10_equalized_four_wheel_trend.png')
    # 11 — distribution percentages
    fig,ax=plt.subplots(figsize=(11,4.8)); ax.plot(x,[r['front_pct'] for r in rs],label='Front %'); ax.plot(x,[r['rear_pct'] for r in rs],label='Rear %'); ax.plot(x,[r['left_pct'] for r in rs],label='Left %',linestyle='--'); ax.plot(x,[r['right_pct'] for r in rs],label='Right %',linestyle='--'); ax.set(title='Overall Load Distribution Percentages',xlabel='Time / Row',ylabel='Distribution [%]'); ax.grid(alpha=.25); ax.legend(ncol=4); save(fig,'11_distribution_percentages.png')
    # 12 — status timeline encoded as 0/1/2
    def sev(r): return {'SAFE':0,'WARNING':1,'DANGER':2}.get(r['post_status'],0)
    fig,ax=plt.subplots(figsize=(11,3.8)); ax.step(x,[sev(r) for r in rs],where='mid'); ax.set(title='Post-Equalization Status Timeline',xlabel='Time / Row',ylabel='Status'); ax.set_yticks([0,1,2]); ax.set_yticklabels(['SAFE','WARNING','DANGER']); ax.grid(alpha=.25); save(fig,'12_status_timeline.png')
    # 13 — transparent rule-based health risk timeline
    fig,ax=plt.subplots(figsize=(11,3.8)); ax.plot(x,[float(r.get('health_risk_score',0)) for r in rs],linewidth=2); ax.axhline(25,linestyle='--',label='WARNING threshold'); ax.axhline(60,linestyle='--',label='DANGER threshold'); ax.set(title='Vehicle Health Risk Timeline',xlabel='Time / Row',ylabel='Screening Risk [0–100]'); ax.set_ylim(0,100); ax.grid(alpha=.25); ax.legend(); save(fig,'13_health_risk_timeline.png')
    return paths

def _styles():
    s=getSampleStyleSheet();
    s.add(ParagraphStyle(name='HLTitle',parent=s['Title'],fontSize=20,textColor=colors.HexColor('#0a6874'),alignment=TA_CENTER,leading=23));
    s.add(ParagraphStyle(name='HLH',parent=s['Heading2'],fontSize=13,textColor=colors.HexColor('#087b88'),spaceBefore=10,spaceAfter=6));
    s.add(ParagraphStyle(name='SectionIntro',parent=s['BodyText'],fontSize=9.5,textColor=colors.HexColor('#42565b'),leading=13));
    s.add(ParagraphStyle(name='Small',parent=s['BodyText'],fontSize=8.2,leading=10,textColor=colors.HexColor('#25373b')));
    s.add(ParagraphStyle(name='StatusGreen',parent=s['BodyText'],fontSize=8.2,textColor=colors.HexColor('#07864f'),leading=10));
    s.add(ParagraphStyle(name='StatusRed',parent=s['BodyText'],fontSize=8.2,textColor=colors.HexColor('#c51f35'),leading=10));
    s.add(ParagraphStyle(name='StatusAmber',parent=s['BodyText'],fontSize=8.2,textColor=colors.HexColor('#9a6800'),leading=10));
    return s

def _report_footer(canvas, doc):
    canvas.saveState()
    canvas.setStrokeColor(colors.HexColor('#d7e3e5'))
    canvas.line(30, 24, A4[0] - 30, 24)
    canvas.setFont('Helvetica', 7)
    canvas.setFillColor(colors.HexColor('#64777b'))
    canvas.drawString(30, 13, 'HydroLevel · Vehicle Load Intelligence & Digital Twin · Engineering screening report')
    canvas.drawRightString(A4[0] - 30, 13, f'Page {doc.page}')
    canvas.restoreState()

def _status_style(status):
    status = str(status).upper()
    if status == 'SAFE': return colors.HexColor('#e7f9ef'), colors.HexColor('#07864f')
    if status == 'DANGER': return colors.HexColor('#ffe9ed'), colors.HexColor('#c51f35')
    return colors.HexColor('#fff5dc'), colors.HexColor('#9a6800')

def build_pdf(payload,path,charts):
    if payload['count']<20: raise ValueError('Minimum 20 valid rows required for full engineering report export.')
    root=Path(__file__).resolve().parents[2]; styles=_styles();
    doc=SimpleDocTemplate(str(path),pagesize=A4,rightMargin=30,leftMargin=30,topMargin=28,bottomMargin=36); story=[]
    logo=root/'frontend/assets/hydrolevel-logo.png'; team=root/'frontend/assets/volts-and-bolts-logo.png'; group=root/'frontend/assets/team-group.png'
    header=[]
    if logo.exists(): header.append(Image(str(logo),width=56,height=56))
    header.append(Paragraph('HYDROLEVEL<br/><font size="9">Vehicle Load Intelligence & Digital Twin</font>',styles['HLTitle']))
    if team.exists(): header.append(Image(str(team),width=56,height=56))
    h=Table([header],colWidths=[70,350,70]); h.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(0,0),(-1,-1),'CENTER'),('BOX',(0,0),(-1,-1),.5,colors.HexColor('#9db7ba'))])); story += [h,Spacer(1,8)]
    story += [Paragraph('Engineering Analysis Report',styles['HLTitle']),Paragraph('Prototype / Academic Engineering Project — Team Volts and Bolts',styles['Normal']),Spacer(1,8)]
    sm=payload['summary'];
    info=[['Parameter','Value'],['Completed rows included',payload.get('completed_count',sm['rows'])],['Valid rows in exported package',sm['rows']],['Minimum rows required',20],['Measurement points','FL · FR · RL · RR'],['Screening threshold',f"±{sm['threshold_kg']:.2f} kg"],['Equalization blend',f"{sm['equalization_blend']*100:.0f}% toward arithmetic reference"],['Mean total load',f"{sm['total_load_mean']:.2f} kg"],['Pre-alert rows',sm['pre_abnormal_rows']],['Post-alert rows',sm['post_abnormal_rows']]]
    t=Table(info,colWidths=[240,280]); t.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#167f8b')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),.3,colors.HexColor('#a9bdc0')),('FONTSIZE',(0,0),(-1,-1),8)])); story += [t,Spacer(1,8)]
    story += [Paragraph('Executive Summary',styles['HLH']),Paragraph(payload.get('ai_overall',''),styles['BodyText']),Paragraph('Data provenance: imported/user-provided measurements are retained as RAW values; derived values are calculated by HydroLevel. Demo/simulated data must be identified as such outside this report.',styles['Small'])]
    health = sm.get('health') or {}
    health_rows = [['Vehicle Health Indicator','Result'],['Risk level',str(health.get('risk_level','SAFE'))],['Screening risk score',f"{float(health.get('risk_score',0)):.1f} / 100"],['Historical rows',str(health.get('history_rows',sm.get('rows',0)))],['Alert rate',f"{float(health.get('alert_rate_percent',0)):.1f}%"],['Deviation trend',str((health.get('trend') or {}).get('deviation_direction','stable/decreasing'))],['Early-warning recommendation',str(health.get('recommendation','Continue monitoring.'))],['Indicators','; '.join(health.get('indicators') or ['None'])]]
    ht=Table(health_rows,colWidths=[180,340],repeatRows=1); ht.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#167f8b')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),.3,colors.HexColor('#a9bdc0')),('FONTSIZE',(0,0),(-1,-1),8),('VALIGN',(0,0),(-1,-1),'TOP')])); story += [Paragraph('Vehicle Health & Early Warning',styles['HLH']),ht,Spacer(1,8)]
    review=payload.get('insurance_review') or payload.get('summary',{}).get('insurance_review',{})
    story += [Paragraph('Insurance Review Support',styles['HLH'])]
    review_data=[
        ['Item','HydroLevel result'],
        ['Review status', Paragraph(str(review.get('status','WAITING')), styles['Small'])],
        ['Engineering screening', Paragraph(str(review.get('engineering_screening','NOT AVAILABLE')), styles['Small'])],
        ['Post-alert rate',f"{float(review.get('post_abnormal_rate_percent',0)):.1f}%"],
        ['Engineering insurance outcome', str(review.get('payout_decision','ENGINEERING SCREENING ONLY'))],
        ['Manual review','REQUIRED'],
        ['Scope',Paragraph(str(review.get('note','Engineering evidence only.')), styles['Small'])]
    ]
    rt=Table(review_data,colWidths=[170,350], repeatRows=1); rt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#167f8b')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),.3,colors.HexColor('#a9bdc0')),('FONTSIZE',(0,0),(-1,-1),8),('VALIGN',(0,0),(-1,-1),'TOP')])); story += [rt,Spacer(1,8)]
    story += [Paragraph('Vehicle / Test Details',styles['HLH'])]
    vehicle_rows=[['Detail','Value']]
    for key, value in (payload.get('vehicle_details') or {}).items():
        vehicle_rows.append([str(key).replace('_',' ').title(), str(value)])
    vehicle_rows += [
        ['Completed rows', str(payload.get('completed_count', sm.get('rows', 0)))],
        ['Mean total load', f"{sm.get('total_load_mean',0):.2f} kg"],
        ['Minimum total load', f"{sm.get('total_load_min',0):.2f} kg"],
        ['Maximum total load', f"{sm.get('total_load_max',0):.2f} kg"],
        ['Maximum wheel load', f"{sm.get('max_wheel_load',0):.2f} kg"],
        ['Minimum wheel load', f"{sm.get('min_wheel_load',0):.2f} kg"],
        ['Front / rear mean distribution', f"{sm.get('front_pct_mean',0):.2f}% / {sm.get('rear_pct_mean',0):.2f}%"],
        ['Left / right mean distribution', f"{sm.get('left_pct_mean',0):.2f}% / {sm.get('right_pct_mean',0):.2f}%"],
    ]
    vt=Table(vehicle_rows,colWidths=[240,280]); vt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#167f8b')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),.3,colors.HexColor('#a9bdc0')),('FONTSIZE',(0,0),(-1,-1),8),('VALIGN',(0,0),(-1,-1),'TOP')])); story += [vt,Spacer(1,8)]
    story += [Paragraph('Import / Validation Log',styles['HLH'])]
    errors = payload.get('validation_errors') or []
    story += [Paragraph('<br/>'.join(str(e) for e in errors[:80]) if errors else 'No validation errors were reported during import.',styles['Small']),Spacer(1,8)]
    story += [PageBreak(),Paragraph('Row-by-Row Analysis',styles['HLH'])]
    data=[['Row','FL raw','FR raw','RL raw','RR raw','FL eq','FR eq','RL eq','RR eq','Total','Avg','Pre','Post','Alerts']]
    for r in payload['results']: data.append([r['index']]+[f"{r['raw'][p]:.1f}" for p in POS]+[f"{r['equalized'][p]:.1f}" for p in POS]+[f"{r['total']:.1f}",f"{r['average']:.1f}",r['pre_status'],r['post_status'],','.join(r['pre_alerts']) or '—'])
    for start in range(0,len(data),28):
        chunk=data[start:start+28]
        t=Table(chunk,repeatRows=1,colWidths=[22,36,36,36,36,36,36,36,36,44,40,40,40,64])
        row_styles=[('BACKGROUND',(0,0),(-1,0),colors.HexColor('#087b88')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),.25,colors.HexColor('#b2c2c4')),('FONTSIZE',(0,0),(-1,-1),6.2),('VALIGN',(0,0),(-1,-1),'MIDDLE')]
        for local_idx, row in enumerate(chunk[1:], start=1):
            post=str(row[12]).upper()
            bg, fg=_status_style(post)
            row_styles.append(('BACKGROUND',(12,local_idx),(12,local_idx),bg))
            row_styles.append(('TEXTCOLOR',(12,local_idx),(12,local_idx),fg))
        t.setStyle(TableStyle(row_styles))
        story += [t,Spacer(1,6)]
        if start+28<len(data): story.append(PageBreak())
    story += [PageBreak(),Paragraph('Engineering Graph Package',styles['HLH']),Paragraph('Graphs are generated from the same processed dataset used by the dashboard. Raw and equalized values are kept separate.',styles['Small'])]
    for i,ch in enumerate(charts):
        story += [Paragraph(Path(ch).stem.replace('_',' ').upper(),styles['Normal']),Image(str(ch),width=510,height=230 if i!=8 else 260)]
        if i<len(charts)-1: story.append(PageBreak())
    story += [PageBreak(),Paragraph('HydroAI Engineering Insight',styles['HLH']),Paragraph(payload.get('ai_overall',''),styles['BodyText']),Spacer(1,8),Paragraph('Engineering note: the ±10 kg value is a project screening threshold. Manufacturer specifications, certified test procedures and applicable regulations remain authoritative.',styles['Small'])]

    # Team appendix is intentionally at the end of the report.
    story += [PageBreak(),Paragraph('Project Team & Contact',styles['HLH'])]
    if group.exists():
        story += [Image(str(group),width=510,height=170),Spacer(1,8)]
    team_data=[[Paragraph('<b>Member</b>',styles['Small']),Paragraph('<b>Role</b>',styles['Small']),Paragraph('<b>Responsibility</b>',styles['Small']),Paragraph('<b>Contact</b>',styles['Small'])]]
    for n,role,edu,res,email,phone,li in TEAM:
        team_data.append([Paragraph(n,styles['Small']),Paragraph(role,styles['Small']),Paragraph(res,styles['Small']),Paragraph(f'{email}<br/>{phone}',styles['Small'])])
    tt=Table(team_data,colWidths=[90,125,190,115],repeatRows=1)
    tt.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),colors.HexColor('#167f8b')),('TEXTCOLOR',(0,0),(-1,0),colors.white),('GRID',(0,0),(-1,-1),.25,colors.HexColor('#b2c2c4')),('VALIGN',(0,0),(-1,-1),'TOP'),('FONTSIZE',(0,0),(-1,-1),7)]))
    story += [tt,Spacer(1,10)]
    photo_files=[('Saai Varshan S','saai-varshan.png'),('Suheerthan S','suheerthan.png'),('Santhosh R','santhosh.png')]
    cells=[]
    for name,fn in photo_files:
        fp=root/f'frontend/assets/{fn}'
        if fp.exists(): cells.append([Image(str(fp),width=82,height=82),Paragraph(f'<b>{name}</b>',styles['Small'])])
    if cells:
        pt=Table([cells],colWidths=[170,170,170]); pt.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE'),('ALIGN',(0,0),(-1,-1),'CENTER'),('GRID',(0,0),(-1,-1),.25,colors.HexColor('#b2c2c4'))])); story += [pt]
    doc.build(story, onFirstPage=_report_footer, onLaterPages=_report_footer)

def build_xlsx(payload,path):
    if payload['count']<20: raise ValueError('Minimum 20 valid rows required for full engineering report export.')
    wb=Workbook(); s=wb.active; s.title='Summary';
    s.append(['HYDROLEVEL ENGINEERING REPORT'])
    s.append(['Completed rows included', payload.get('completed_count', payload['count'])])
    s.append([])
    s.append(['Metric','Value'])
    # Excel cells cannot store Python dictionaries/lists.  Keep the summary
    # human-readable instead of failing when the insurance review object is
    # included in the analysis summary.
    for key, value in payload['summary'].items():
        if isinstance(value, (dict, list, tuple)):
            value = json.dumps(value, ensure_ascii=False)
        s.append([key, value])
    raw=wb.create_sheet('Raw Data'); raw.append(['Row','FL','FR','RL','RR']); proc=wb.create_sheet('Processed Data'); proc.append(['Row','Total','Average','Front Axle','Rear Axle','Left','Right','Front %','Rear %','Left %','Right %','Pre Status','Post Status','Health Risk','Health Level','CG X','CG Y','FL Eq','FR Eq','RL Eq','RR Eq']); ev=wb.create_sheet('Events'); ev.append(['Row','Position','Raw','Deviation','Equalized','Eq Deviation','Pre Alert','Post Alert'])
    for r in payload['results']:
        raw.append([r['index']]+[r['raw'][p] for p in POS]); proc.append([r['index'],r['total'],r['average'],r['front_axle'],r['rear_axle'],r['left_side'],r['right_side'],r['front_pct'],r['rear_pct'],r['left_pct'],r['right_pct'],r['pre_status'],r['post_status'],r['cg_x'],r['cg_y']]+[r['equalized'][p] for p in POS])
        for p in POS:
            if p in r['pre_alerts'] or p in r['post_alerts']: ev.append([r['index'],p,r['raw'][p],r['deviations'][p],r['equalized'][p],r['equalized_deviations'][p],p in r['pre_alerts'],p in r['post_alerts']])
    review=payload.get('insurance_review') or payload.get('summary',{}).get('insurance_review',{})
    ins=wb.create_sheet('Insurance Review')
    ins_rows=[
        ['HYDROLEVEL INSURANCE REVIEW SUPPORT'],
        ['Status', review.get('status','WAITING')],
        ['Engineering screening', review.get('engineering_screening','NOT AVAILABLE')],
        ['Post-alert rate %', float(review.get('post_abnormal_rate_percent',0))],
        ['Engineering insurance outcome', review.get('payout_decision','ENGINEERING SCREENING ONLY')],
        ['Manual review required', 'YES'],
        ['Note', review.get('note','Engineering evidence only.')],
    ]
    for row in ins_rows: ins.append(row)

    vd=wb.create_sheet('Vehicle Details')
    vd.append(['VEHICLE / TEST DETAIL','VALUE'])
    for key, value in (payload.get('vehicle_details') or {}).items():
        vd.append([str(key).replace('_',' ').title(), value])
    vd.append(['Completed rows', payload.get('completed_count', payload['count'])])
    vd.append(['Mean total load kg', payload['summary'].get('total_load_mean', 0)])
    vd.append(['Maximum wheel load kg', payload['summary'].get('max_wheel_load', 0)])
    vd.append(['Post-alert rate %', payload['summary'].get('post_abnormal_rate_percent', 0)])

    er=wb.create_sheet('Validation Errors')
    er.append(['IMPORT / VALIDATION ERROR'])
    errors=payload.get('validation_errors') or []
    if errors:
        for item in errors: er.append([item])
    else:
        er.append(['No validation errors were reported during import.'])

    for sh in wb.worksheets:
        for c in sh[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='167F8B')
        sh.freeze_panes='A2'; sh.auto_filter.ref=sh.dimensions
    wb.save(path)

def build_csv(payload,path):
    if payload['count']<20: raise ValueError('Minimum 20 valid rows required for full engineering report export.')
    with open(path,'w',newline='',encoding='utf-8') as f:
        w=csv.writer(f); w.writerow(['Row','FL_raw','FR_raw','RL_raw','RR_raw','Total','Average','FL_eq','FR_eq','RL_eq','RR_eq','Pre_Status','Post_Status','CG_X','CG_Y'])
        for r in payload['results']: w.writerow([r['index']]+[r['raw'][p] for p in POS]+[r['total'],r['average']]+[r['equalized'][p] for p in POS]+[r['pre_status'],r['post_status'],r['cg_x'],r['cg_y']])
